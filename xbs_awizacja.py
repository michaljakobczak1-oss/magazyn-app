"""Wypełnianie formularza awizacyjnego XBS (Excel) na podstawie pozycji WZ."""
from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

TEMPLATE_PATH = Path(__file__).resolve().parent / "forms" / "xbs_awizacja.xlsx"
PRODUCT_START = 22
PRODUCT_SLOTS = 3  # wiersze wzorcowe w szablonie
MATERIAL_OPTIONS = ("Papier", "Tekstylia", "Spożywcze")


def _copy_row_style(ws, src_row, dst_row, max_col=12):
    for col in range(1, max_col + 1):
        s = ws.cell(src_row, col)
        d = ws.cell(dst_row, col)
        if s.has_style:
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy(s.protection)
            d.alignment = copy(s.alignment)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _fix_product_row_frames(ws, start_row, end_row):
    """Po rozscaleniu J22:J23 (i insert_rows) przywraca ramki wierszy produktowych."""
    template = PRODUCT_START
    for row in range(start_row, end_row + 1):
        if row == template:
            continue
        _copy_row_style(ws, template, row, max_col=10)
        _ensure_name_merge(ws, row)


def _unmerge_exact(ws, min_row, max_row, min_col=None, max_col=None):
    to_remove = []
    for m in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = range_boundaries(str(m))
        if r2 < min_row or r1 > max_row:
            continue
        if min_col is not None and (c2 < min_col or c1 > max_col):
            continue
        to_remove.append(str(m))
    for m in to_remove:
        ws.unmerge_cells(m)


def _set_top_left(ws, coord, value):
    """Ustaw wartość także gdy komórka jest częścią scalenia."""
    cell = ws[coord]
    for m in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(m))
        if min_r <= cell.row <= max_r and min_c <= cell.column <= max_c:
            ws.cell(min_r, min_c).value = value
            return
    cell.value = value


def _ensure_name_merge(ws, row):
    """Scal Nazwa C:D dla wiersza produktu (jak w szablonie)."""
    _unmerge_exact(ws, row, row, min_col=3, max_col=4)
    try:
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    except ValueError:
        pass


def _clear_product_row(ws, row):
    """Czyści wartości w wierszu produktu bez niszczenia układu kolumn."""
    for col in (2, 3, 5, 6, 7, 8, 9, 10):  # B,C,E,F,G,H,I,J
        _set_top_left(ws, ws.cell(row, col).coordinate, None)


def _maybe_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        num = float(s)
        return int(num) if num.is_integer() else num
    except ValueError:
        return s


def build_xbs_awizacja_xlsx(items, meta):
    """Zwraca BytesIO z wypełnionym formularzem.

    items: lista dict z kluczami code, name, quantity
    meta: pola awizacji + qty_per_pallet, weight, material, pallets
    """
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Brak szablonu XBS: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    supplier = (meta.get("supplier") or "").strip()
    supplier_person = (meta.get("supplier_person") or "").strip()
    supplier_phone = (meta.get("supplier_phone") or "").strip()
    supplier_address = (meta.get("supplier_address") or "").strip()
    order_no = (meta.get("order_no") or "").strip()
    delivery_date = (meta.get("delivery_date") or "").strip()
    delivery_time = (meta.get("delivery_time") or "").strip()
    notes = (meta.get("notes") or "").strip()
    carrier = (meta.get("carrier") or "").strip()
    plate = (meta.get("plate") or "").strip()
    material = (meta.get("material") or "").strip()
    if material not in MATERIAL_OPTIONS:
        material = ""
    per_item = meta.get("items") or {}
    default_qty_per_pallet = meta.get("qty_per_pallet")
    default_weight = meta.get("weight")
    default_material = material
    default_pallets = meta.get("pallets")

    _set_top_left(ws, "D14", supplier)
    _set_top_left(ws, "D15", supplier_person)
    _set_top_left(ws, "D16", supplier_phone)
    _set_top_left(ws, "D17", supplier_address)
    _set_top_left(ws, "G14", order_no)

    # Szablon ma J22:J23 – rozłącz, żeby każdy wiersz miał własną „Suma palet”
    _unmerge_exact(ws, PRODUCT_START, PRODUCT_START + PRODUCT_SLOTS - 1, min_col=10, max_col=10)
    _fix_product_row_frames(ws, PRODUCT_START + 1, PRODUCT_START + PRODUCT_SLOTS - 1)

    # Wyczyść 3 sloty produktowe (bez wywalania scalenia Nazwa)
    for r in range(PRODUCT_START, PRODUCT_START + PRODUCT_SLOTS):
        _ensure_name_merge(ws, r)
        _clear_product_row(ws, r)

    n = len(items)
    if n > PRODUCT_SLOTS:
        extra = n - PRODUCT_SLOTS
        insert_at = PRODUCT_START + PRODUCT_SLOTS
        ws.insert_rows(insert_at, amount=extra)
        for i in range(extra):
            row = insert_at + i
            _copy_row_style(ws, PRODUCT_START, row, max_col=10)
            _ensure_name_merge(ws, row)

    last_product_row = PRODUCT_START + max(n, PRODUCT_SLOTS) - 1
    _fix_product_row_frames(ws, PRODUCT_START + 1, last_product_row)

    for i, it in enumerate(items):
        row = PRODUCT_START + i
        _ensure_name_merge(ws, row)
        ws.cell(row, 2).value = i + 1  # LP
        _set_top_left(ws, ws.cell(row, 3).coordinate, (it.get("name") or "").strip())
        ws.cell(row, 5).value = (it.get("code") or "").strip()
        qty = it.get("quantity")
        try:
            ws.cell(row, 6).value = int(qty)
        except (TypeError, ValueError):
            ws.cell(row, 6).value = qty
        rid = str(it.get("rid") or "")
        row_meta = per_item.get(rid) or {}
        qty_per_pallet = _maybe_number(
            row_meta.get("qty_per_pallet") or default_qty_per_pallet
        )
        weight = _maybe_number(row_meta.get("weight") or default_weight)
        row_material = (row_meta.get("material") or default_material or "").strip()
        if row_material not in MATERIAL_OPTIONS:
            row_material = ""
        pallets = _maybe_number(row_meta.get("pallets") or default_pallets)
        if qty_per_pallet is not None:
            ws.cell(row, 7).value = qty_per_pallet
        if weight is not None:
            ws.cell(row, 8).value = weight
        if row_material:
            ws.cell(row, 9).value = row_material
        if pallets is not None:
            ws.cell(row, 10).value = pallets

    # Etykiety poniżej produktów (po ewentualnym insert_rows)
    label_map = {}
    for r in range(1, (ws.max_row or 80) + 1):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                key = v.strip().lower()
                if key in ("data dostawy", "godz.dostawy", "uwagi:", "przewoźnik",
                           "nr rej.samochodu"):
                    label_map[key] = (r, c)

    def _fill_after_label(label, value, col_offset=1):
        pos = label_map.get(label)
        if not pos:
            return
        r, c = pos
        _set_top_left(ws, ws.cell(r, c + col_offset).coordinate, value)

    parsed_date = None
    if delivery_date:
        try:
            parsed_date = datetime.strptime(delivery_date, "%Y-%m-%d")
        except ValueError:
            parsed_date = delivery_date
    _fill_after_label("data dostawy", parsed_date)
    _fill_after_label("godz.dostawy", delivery_time)
    _fill_after_label("uwagi:", notes)
    _fill_after_label("przewoźnik", carrier)
    _fill_after_label("nr rej.samochodu", plate)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def xbs_filename(stamp=None):
    stamp = stamp or datetime.now().strftime("%Y%m%d_%H%M")
    return f"XBS_Awizacja_{stamp}.xlsx"
