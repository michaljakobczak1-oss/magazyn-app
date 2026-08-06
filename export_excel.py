"""
Eksport katalogu sprzętu:
- Excel z osadzonymi miniaturami (PM),
- Excel w formacie importu + ZIP zdjęć (admin).
"""
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage

HEADERS = [
    "Kod",
    "Miniatura",
    "Nazwa produktu",
    "Nr projektu",
    "Wymiary",
    "Magazyn",
    "Miejsce",
    "Własność",
    "Brand",
    "Rodzaj materiału",
    "Stan techn.",
    "Ilość",
]

IMPORT_HEADERS = [
    "Kod",
    "Nazwa produktu",
    "Numer projektu",
    "Wymiary",
    "Magazyn",
    "Miejsce w magazynie",
    "Własność",
    "Brand",
    "Rodzaj materiału",
    "Stan techniczny",
    "Stan magazynowy (szt.)",
    "Składowanie / pakowanie",
    "Plik zdjęcia",
]

THUMB_W = 95
THUMB_H = 110
ROW_HEIGHT = 82
COL_WIDTHS = {
    "A": 12, "B": 17, "C": 36, "D": 14, "E": 16,
    "F": 14, "G": 16, "H": 16, "I": 14, "J": 16,
    "K": 12, "L": 10,
}


def material_label(material_type):
    if (material_type or "").strip().lower() == "wlasny":
        return "materiał własny"
    return "materiał klienta"


def _make_thumb_bytes(path: Path):
    """Zwraca BytesIO JPEG albo None."""
    try:
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            im.thumbnail((THUMB_W, THUMB_H), PILImage.Resampling.LANCZOS)
            canvas = PILImage.new("RGB", (THUMB_W, THUMB_H), (255, 255, 255))
            ox = (THUMB_W - im.width) // 2
            oy = (THUMB_H - im.height) // 2
            canvas.paste(im, (ox, oy))
            buf = BytesIO()
            canvas.save(buf, format="JPEG", quality=85)
            buf.seek(0)
            return buf
    except OSError:
        return None


def build_catalog_miniatures_xlsx(rows, upload_dir):
    """
    rows: iterable of mapping-like objects with keys:
      code, name, project_number, dimensions, warehouse_name, location,
      owner, brand, material_type, condition, quantity, storage_instructions, photo
    upload_dir: Path do static/uploads
    Zwraca BytesIO z plikiem .xlsx.
    """
    upload_dir = Path(upload_dir)
    wb = Workbook()
    ws = wb.active
    ws.title = "Import (miniatury)"

    header_font = Font(bold=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    keep_alive = []

    for i, row in enumerate(rows, start=2):
        values = [
            row["code"],
            None,
            row["name"],
            row["project_number"] or "",
            row["dimensions"] or "",
            row["warehouse_name"] or "",
            row["location"] or "",
            row["owner"] or "",
            row["brand"] or "",
            material_label(row["material_type"]),
            row["condition"] or "sprawny",
            row["quantity"] if row["quantity"] is not None else "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = ROW_HEIGHT

        photo_name = (row["photo"] or "").strip() if row["photo"] else ""
        if photo_name:
            path = upload_dir / Path(photo_name).name
            if path.is_file():
                thumb = _make_thumb_bytes(path)
                if thumb is not None:
                    keep_alive.append(thumb)
                    xl_img = XLImage(thumb)
                    xl_img.width = THUMB_W
                    xl_img.height = THUMB_H
                    ws.add_image(xl_img, f"B{i}")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    del keep_alive
    return out


def build_catalog_import_xlsx(rows):
    """
    Excel jak przy imporcie (arkusz Import, bez miniatur).
    rows: mapping z kluczami DB + warehouse_name + photo_file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    header_font = Font(bold=True)
    for col, title in enumerate(IMPORT_HEADERS, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
    for i, row in enumerate(rows, start=2):
        values = [
            row["code"],
            row["name"],
            row.get("project_number") or "",
            row.get("dimensions") or "",
            row.get("warehouse_name") or "",
            row.get("location") or "",
            row.get("owner") or "",
            row.get("brand") or "",
            material_label(row.get("material_type")),
            row.get("condition") or "sprawny",
            row.get("quantity") if row.get("quantity") is not None else "",
            row.get("storage_instructions") or "",
            row.get("photo_file") or "",
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(i, col, val)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_photo_zip_names(code, source_filenames, upload_dir):
    """
    Mapuje pliki z uploads na nazwy w stylu importu: KOD.ext, KOD_2.ext, …
    Zwraca (lista nazw do kolumny Plik zdjęcia, lista (src Path, arcname w ZIP)).
    """
    upload_dir = Path(upload_dir)
    names = []
    copies = []
    used = set()
    idx = 0
    for raw in source_filenames:
        if not raw:
            continue
        src = upload_dir / Path(str(raw)).name
        if not src.is_file():
            continue
        ext = src.suffix.lower() if src.suffix else ".jpeg"
        if ext not in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            ext = ".jpeg"
        idx += 1
        if idx == 1:
            candidate = f"{code}{ext}"
        else:
            candidate = f"{code}_{idx}{ext}"
        base = candidate
        n = 1
        while candidate.lower() in used:
            n += 1
            stem = Path(base).stem
            candidate = f"{stem}_{n}{ext}"
        used.add(candidate.lower())
        names.append(candidate)
        copies.append((src, f"zdjecia/{candidate}"))
    return names, copies
