"""
Import rejestru sprzętu z arkusza Excel.

Użycie:
    python import_excel.py plik.xlsx --photos katalog_ze_zdjeciami [--sheet Import] [--dry-run]
    python import_excel.py plik.xlsx --photos zdjecia --update   # nadpisz istniejące kody

Oczekiwane kolumny (nagłówki, wielkość liter bez znaczenia, kolejność dowolna):
    Kod | Nazwa / Nazwa produktu | Numer projektu | Wymiary
    Magazyn | Miejsce w magazynie | Własność | Brand
    Rodzaj materiału | Stan techniczny | Stan magazynowy (szt.) / Ilość
    Składowanie / pakowanie | Plik zdjęcia / Zdjęcie

Zdjęcia:
    1) kolumna „Plik zdjęcia” – nazwy plików z katalogu --photos (wiele po przecinku),
    2) fallback: plik nazwany kodem (np. CC0043.png),
    3) zdjęcia osadzone w xlsx (po wierszu).

Wymaga: pip install openpyxl
"""
import argparse
import re
import shutil
import sys
import unicodedata
import uuid
from pathlib import Path

from openpyxl import load_workbook

from db import get_db, init_db

BASE = Path(__file__).parent
UPLOAD_DIR = BASE / "static" / "uploads"
ALLOWED_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp"}

# nagłówek w arkuszu -> pole w bazie
HEADER_MAP = {
    "kod": "code",
    "nowy kod": "code",
    "zdjęcie": "photo_file",
    "zdjecie": "photo_file",
    "foto": "photo_file",
    "plik zdjęcia": "photo_file",
    "plik zdjecia": "photo_file",
    "nazwa": "name",
    "nazwa produktu": "name",
    "ilość": "quantity",
    "ilosc": "quantity",
    "szt": "quantity",
    "stan magazynowy (szt.)": "quantity",
    "stan magazynowy": "quantity",
    "rozmiar/waga/opis": "dimensions",
    "rozmiar": "dimensions",
    "wymiary": "dimensions",
    "opis": "dimensions",
    "miejsce": "location",
    "miejsce w magazynie": "location",
    "magazyn": "warehouse",
    "własność": "owner",
    "wlasnosc": "owner",
    "projekt": "project_number",
    "numer projektu": "project_number",
    "nr projektu": "project_number",
    "brand": "brand",
    "marka": "brand",
    "rodzaj materiału": "material_type",
    "rodzaj materialu": "material_type",
    "materiał": "material_type",
    "stan techniczny": "condition",
    "stan": "condition",
    "składowanie / pakowanie": "storage_instructions",
    "skladowanie / pakowanie": "storage_instructions",
    "składowanie": "storage_instructions",
    "pakowanie": "storage_instructions",
}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def nfc(s):
    return unicodedata.normalize("NFC", str(s or ""))


def strip_diacritics(s):
    """HGOŚW → HGOSW (dopasowanie NFC/NFD i różnic po rozpakowaniu ZIP na Linuxie)."""
    decomposed = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in decomposed if unicodedata.category(c) != "Mn")


def photo_lookup_keys(filename):
    """Wszystkie warianty klucza do indeksu zdjęć."""
    name = Path(str(filename or "").strip()).name
    if not name:
        return []
    keys = []
    seen = set()
    for form in (
        name,
        unicodedata.normalize("NFC", name),
        unicodedata.normalize("NFD", name),
        strip_diacritics(name),
        strip_diacritics(unicodedata.normalize("NFC", name)),
    ):
        k = form.casefold()
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


def map_headers(ws):
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = HEADER_MAP.get(norm(cell.value))
        if key and key not in headers:
            headers[key] = idx
    missing = {"code", "name"} - set(headers)
    if missing:
        raise ValueError(
            f"Brak wymaganych kolumn w arkuszu: {', '.join(sorted(missing))}. "
            f"Znalezione nagłówki: {[c.value for c in ws[1]]}"
        )
    return headers


def extract_embedded_images(ws):
    """Zwraca {nr_wiersza: bytes} dla zdjęć osadzonych w arkuszu."""
    out = {}
    for img in getattr(ws, "_images", []):
        try:
            row = img.anchor._from.row + 1  # 0-indexed
            out[row] = img._data()
        except Exception:
            pass
    return out


def build_photo_index(photos_dir):
    """Mapa wariantów nazwy (NFC/NFD/bez diakrytyków) -> Path. Szuka też w podfolderach."""
    if not photos_dir or not photos_dir.exists():
        return {}
    idx = {}
    files = []
    root = Path(photos_dir)
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if "__MACOSX" in p.parts:
            continue
        if p.suffix.lower() not in ALLOWED_EXT:
            continue
        files.append(p)
    # jeśli rglob nic nie dał, spróbuj tylko poziom root
    if not files:
        files = [
            p for p in root.iterdir()
            if p.is_file() and p.suffix.lower() in ALLOWED_EXT
        ]
    for p in files:
        for key in photo_lookup_keys(p.name):
            idx.setdefault(key, p)
    return idx


def find_photo_by_name(photo_index, filename):
    if not filename or not photo_index:
        return None
    for key in photo_lookup_keys(filename):
        if key in photo_index:
            return photo_index[key]
    # bez rozszerzenia – spróbuj znanych
    stem = Path(str(filename).strip()).stem
    for ext in ALLOWED_EXT:
        for key in photo_lookup_keys(stem + ext):
            if key in photo_index:
                return photo_index[key]
    return None


def find_photo_by_code(photo_index, code):
    if not code:
        return None
    for candidate in (code, strip_diacritics(code), code.lower(), code.upper()):
        for ext in ALLOWED_EXT:
            p = find_photo_by_name(photo_index, candidate + ext)
            if p:
                return p
    return None


def parse_photo_names(raw):
    if not raw:
        return []
    parts = re.split(r"[,;|]", str(raw))
    return [nfc(p).strip() for p in parts if nfc(p).strip()]


def save_photo_bytes(data, ext=".png"):
    fname = f"{uuid.uuid4().hex}{ext}"
    (UPLOAD_DIR / fname).write_bytes(data)
    return fname


def save_photo_file(path):
    fname = f"{uuid.uuid4().hex}{path.suffix.lower()}"
    shutil.copy(path, UPLOAD_DIR / fname)
    return fname


def normalize_material_type(raw, code):
    s = norm(raw)
    if "własn" in s or "wlasn" in s:
        return "wlasny"
    if "klient" in s:
        return "klient"
    return "wlasny" if str(code).startswith("00") else "klient"


def normalize_condition(raw):
    s = norm(raw)
    if not s:
        return "sprawny"
    if "utyliz" in s:
        return "do utylizacji"
    if "uszkod" in s:
        return "uszkodzony"
    if "spraw" in s:
        return "sprawny"
    return "sprawny"


def get_or_create_warehouse(con, name, dry_run=False, cache=None):
    name = (name or "").strip()
    if not name:
        return None
    if cache is not None and name in cache:
        return cache[name]
    row = con.execute("SELECT id FROM warehouses WHERE name=?", (name,)).fetchone()
    if row:
        wid = row["id"]
    elif dry_run:
        wid = f"(new:{name})"
    else:
        cur = con.execute(
            "INSERT INTO warehouses (name, active) VALUES (?,1)", (name,)
        )
        wid = cur.lastrowid
        print(f"Utworzono magazyn: {name}")
    if cache is not None:
        cache[name] = wid
    return wid


def pick_sheet(wb, preferred):
    if preferred:
        if preferred not in wb.sheetnames:
            raise ValueError(f"Brak arkusza {preferred!r}. Dostępne: {wb.sheetnames}")
        return wb[preferred]
    for name in ("Import", "import", "IMPORT"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def run_import(xlsx_path, photos_dir=None, sheet="Import", warehouse=None,
               update=False, dry_run=False, log=None):
    """Uruchamia import. Zwraca dict: added, updated, skipped, no_photo, messages."""
    messages = log if log is not None else []
    init_db()
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    photos_dir = Path(photos_dir) if photos_dir else None
    photo_index = build_photo_index(photos_dir)
    if photos_dir and not photo_index:
        messages.append(f"UWAGA: katalog zdjęć pusty / nie znaleziono plików: {photos_dir}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = pick_sheet(wb, sheet)
    headers = map_headers(ws)
    embedded = extract_embedded_images(ws)
    messages.append(f"Arkusz: {ws.title} | kolumny: {sorted(headers)}")
    messages.append(f"Indeks zdjęć: {len(photo_index)} plików")

    con = get_db()
    wh_cache = {}
    if warehouse:
        get_or_create_warehouse(con, warehouse, dry_run, wh_cache)

    added, updated, skipped, no_photo = 0, 0, 0, 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def val(field):
            col = headers.get(field)
            return row[col - 1].value if col else None

        code = str(val("code") or "").strip()
        name = str(val("name") or "").strip()
        if not code or not name:
            continue

        existing = con.execute(
            "SELECT id FROM equipment WHERE code=?", (code,)
        ).fetchone()
        if existing and not update:
            messages.append(f"POMIJAM (kod istnieje): {code}")
            skipped += 1
            continue

        qty_raw = val("quantity")
        has_qty_col = "quantity" in headers
        if not has_qty_col:
            qty = None  # arkusz bez kolumny ilości (np. Nowe kody) → nie nadpisuj
        elif qty_raw is None or (isinstance(qty_raw, str) and not str(qty_raw).strip()):
            qty = 0
        else:
            try:
                qty = int(float(qty_raw))
                if qty < 0:
                    qty = 0
            except (TypeError, ValueError):
                qty = 0

        wh_name = str(val("warehouse") or "").strip() or (warehouse or "")
        warehouse_id = get_or_create_warehouse(con, wh_name, dry_run, wh_cache) if wh_name else None

        material_type = normalize_material_type(val("material_type"), code) if "material_type" in headers else None
        condition = normalize_condition(val("condition")) if "condition" in headers else None
        storage = str(val("storage_instructions") or "").strip() if "storage_instructions" in headers else None
        location = str(val("location") or "").strip() if "location" in headers else None
        owner = str(val("owner") or "").strip() if "owner" in headers else None
        brand = str(val("brand") or "").strip() if "brand" in headers else None
        project_number = str(val("project_number") or "").strip() if "project_number" in headers else None
        dimensions = str(val("dimensions") or "").strip() if "dimensions" in headers else None

        photo_files = []
        for pname in parse_photo_names(val("photo_file")):
            p = find_photo_by_name(photo_index, pname)
            if p:
                photo_files.append(p)
            else:
                messages.append(f"  brak pliku zdjęcia: {code} → {pname}")

        if not photo_files:
            p = find_photo_by_code(photo_index, code)
            if p:
                photo_files.append(p)

        saved_names = []
        if photo_files:
            if not dry_run:
                for p in photo_files:
                    saved_names.append(save_photo_file(p))
            else:
                saved_names = [p.name for p in photo_files]
        elif r_idx in embedded:
            if not dry_run:
                saved_names = [save_photo_bytes(embedded[r_idx])]
            else:
                saved_names = ["(embedded)"]
        else:
            no_photo += 1

        primary_photo = saved_names[0] if saved_names else None
        qty_display = qty if qty is not None else (0 if not existing else "?")

        if dry_run:
            action = "UPDATE" if existing else "INSERT"
            messages.append(
                f"{action}: {code} | {name} | {qty_display} szt. | {wh_name or '-'} / {location or '-'} | "
                f"zdjęcia: {len(saved_names)}"
            )
            if existing:
                updated += 1
            else:
                added += 1
            continue

        if existing:
            eid = existing["id"]
            # Pola Absent/None w arkuszu (np. „Nowe kody” bez ilości) – nie nadpisuj istniejących
            con.execute(
                """UPDATE equipment SET
                   name=?,
                   project_number=COALESCE(?, project_number),
                   dimensions=COALESCE(?, dimensions),
                   photo=COALESCE(?, photo),
                   location=COALESCE(?, location),
                   warehouse_id=COALESCE(?, warehouse_id),
                   owner=COALESCE(?, owner),
                   brand=COALESCE(?, brand),
                   material_type=COALESCE(?, material_type),
                   condition=COALESCE(?, condition),
                   storage_instructions=COALESCE(?, storage_instructions),
                   quantity=COALESCE(?, quantity)
                   WHERE id=?""",
                (
                    name,
                    project_number if project_number else None,
                    dimensions if dimensions else None,
                    primary_photo,
                    location if location else None,
                    warehouse_id,
                    owner if owner else None,
                    brand if brand else None,
                    material_type,
                    condition,
                    storage if storage else None,
                    qty,
                    eid,
                ),
            )
            if saved_names:
                con.execute("DELETE FROM equipment_photos WHERE equipment_id=?", (eid,))
                for i, fn in enumerate(saved_names):
                    con.execute(
                        "INSERT INTO equipment_photos (equipment_id, filename, sort_order) VALUES (?,?,?)",
                        (eid, fn, i),
                    )
            updated += 1
        else:
            cur = con.execute(
                """INSERT INTO equipment (code, project_number, name, dimensions, photo,
                   location, warehouse_id, owner, brand, material_type,
                   condition, storage_instructions, quantity)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    code,
                    project_number or "",
                    name,
                    dimensions or "",
                    primary_photo,
                    location or "",
                    warehouse_id,
                    owner or "",
                    brand or "",
                    material_type or "klient",
                    condition or "sprawny",
                    storage or "",
                    0 if qty is None else qty,
                ),
            )
            eid = cur.lastrowid
            for i, fn in enumerate(saved_names):
                con.execute(
                    "INSERT INTO equipment_photos (equipment_id, filename, sort_order) VALUES (?,?,?)",
                    (eid, fn, i),
                )
            added += 1

    if not dry_run:
        con.commit()
    con.close()
    summary = (
        f"Gotowe. Dodano: {added}, zaktualizowano: {updated}, "
        f"pominięto: {skipped}, bez zdjęcia: {no_photo}."
    )
    messages.append(summary)
    if dry_run:
        messages.append("(dry-run – nic nie zapisano)")
    return {
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "no_photo": no_photo,
        "messages": messages,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="plik .xlsx z rejestrem")
    ap.add_argument("--photos", help="katalog ze zdjęciami (nazwy jak w kolumnie Plik zdjęcia)")
    ap.add_argument("--sheet", default="Import", help="nazwa arkusza (domyślnie Import)")
    ap.add_argument("--warehouse", help="magazyn domyślny, gdy w wierszu brak Magazynu")
    ap.add_argument("--update", action="store_true",
                    help="aktualizuj istniejące kody zamiast je pomijać")
    ap.add_argument("--dry-run", action="store_true", help="tylko pokaż, nic nie zapisuj")
    args = ap.parse_args()

    try:
        result = run_import(
            args.xlsx,
            photos_dir=args.photos,
            sheet=args.sheet,
            warehouse=args.warehouse,
            update=args.update,
            dry_run=args.dry_run,
        )
    except ValueError as e:
        sys.exit(str(e))
    for line in result["messages"]:
        print(line)


if __name__ == "__main__":
    main()
